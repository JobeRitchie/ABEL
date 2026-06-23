from pathlib import Path

from openpyxl import load_workbook

from abel.models.schemas import (
    CandidateWindow,
    ImportManifest,
    LinkedSession,
    ReviewDecision,
    ReviewDecisionType,
    VideoAsset,
)
from abel.services.export_service import ExportService
from abel.services.import_service import ImportService


def test_export_boutframes_handles_sheet_name_collisions_and_invalid_chars(tmp_path: Path) -> None:
    project_root = tmp_path / "project"
    (project_root / "derived" / "review_tables").mkdir(parents=True)

    # These subjects sanitize/truncate to the same prefix and include invalid Excel title chars.
    subject_a = "Subject:ABCDEF_123456789012345678901"
    subject_b = "Subject/ABCDEF_123456789012345678902"

    manifest = ImportManifest(
        videos=[
            VideoAsset(asset_id="v1", source_path="a.avi", subject_id=subject_a),
            VideoAsset(asset_id="v2", source_path="b.avi", subject_id=subject_b),
        ],
        linked_sessions=[
            LinkedSession(session_id="s1", video_asset_id="v1", pose_asset_id="p1", subject_id=subject_a),
            LinkedSession(session_id="s2", video_asset_id="v2", pose_asset_id="p2", subject_id=subject_b),
        ],
    )
    ImportService().save_manifest(project_root, manifest)

    candidates = [
        CandidateWindow(window_id="c1", session_id="s1", start_frame=10, end_frame=20, behavior_id="rear"),
        CandidateWindow(window_id="c2", session_id="s2", start_frame=30, end_frame=40, behavior_id="rear"),
    ]
    decisions = [
        ReviewDecision(
            decision_id="d1",
            clip_id="c1",
            reviewer="r",
            old_status="unscored",
            new_status="reviewed",
            decision=ReviewDecisionType.ACCEPT,
            adjusted_start_frame=11,
            adjusted_end_frame=19,
        ),
        ReviewDecision(
            decision_id="d2",
            clip_id="c2",
            reviewer="r",
            old_status="unscored",
            new_status="reviewed",
            decision=ReviewDecisionType.ACCEPT,
            adjusted_start_frame=31,
            adjusted_end_frame=39,
        ),
    ]

    service = ExportService()
    service.set_project(project_root)
    out = service.export_boutframes_xlsx(
        candidates,
        decisions,
        filename="boutframes.xlsx",
        include_end_frames=True,
    )

    assert out.success
    assert out.output_path is not None
    wb = load_workbook(out.output_path)

    assert "_bout_counts" in wb.sheetnames
    subject_sheets = [name for name in wb.sheetnames if name != "_bout_counts"]
    assert len(subject_sheets) == 2
    assert len(set(subject_sheets)) == 2

    for sheet_name in subject_sheets:
        assert len(sheet_name) <= 31
        for bad in [":", "\\", "/", "?", "*", "[", "]"]:
            assert bad not in sheet_name

    first_subject_rows = list(wb[subject_sheets[0]].iter_rows(min_row=1, max_row=2, values_only=True))
    assert first_subject_rows[0] == ("rear__start", "rear__end")
