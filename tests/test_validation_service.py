from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
import yaml

from abel.models.schemas import (
    ValidationAnswerRecord,
    ValidationClipRecord,
    ValidationRun,
    ValidationSettings,
)
from abel.services.validation_service import ValidationService


# ---------------------------------------------------------------------------
# Synthetic project fixtures
# ---------------------------------------------------------------------------
def _make_project(tmp_path: Path, *, onset: float = 0.6, min_bout: int = 3, merge_gap: int = 1) -> Path:
    root = tmp_path / "proj"
    (root / "config").mkdir(parents=True)
    (root / "derived" / "temporal_refinement" / "groom").mkdir(parents=True)
    inf_dir = root / "derived" / "inference"
    inf_dir.mkdir(parents=True)

    # One active behavior.
    behaviors = {
        "behaviors": [
            {
                "behavior_id": "groom",
                "name": "Grooming",
                "short_name": "groom",
                "keyboard_shortcut": "g",
                "is_active": True,
            }
        ]
    }
    (root / "config" / "behavior_definitions.yaml").write_text(yaml.safe_dump(behaviors), encoding="utf-8")

    # Threshold settings.
    (root / "config" / "temporal_review_settings.json").write_text(
        json.dumps(
            {
                "__all__": {
                    "onset_threshold": onset,
                    "min_bout_duration_frames": min_bout,
                    "merge_gap_frames": merge_gap,
                }
            }
        ),
        encoding="utf-8",
    )

    # Probability trace: baseline 0.1, a clear bout (10-20 @ 0.9),
    # a near-threshold bout (23-27 @ 0.62).
    probs = [0.1] * 50
    for f in range(10, 21):
        probs[f] = 0.9
    for f in range(23, 28):
        probs[f] = 0.62
    trace = pd.DataFrame({"frame": list(range(50)), "prob_groom": probs})
    trace_path = inf_dir / "session_a__trace.parquet"
    trace.to_parquet(trace_path, index=False)

    (inf_dir / "inference_manifest.json").write_text(
        json.dumps({"trace_paths": {"session_a": str(trace_path)}}), encoding="utf-8"
    )

    (root / "derived" / "temporal_refinement" / "groom" / "latest.json").write_text(
        json.dumps({"inference_dir": str(inf_dir), "postprocess_dir": ""}), encoding="utf-8"
    )
    return root


def _service(root: Path) -> ValidationService:
    svc = ValidationService()
    svc.set_project(root)
    return svc


# ---------------------------------------------------------------------------
# Bout / sampling
# ---------------------------------------------------------------------------
def test_detect_bouts_finds_positive_and_fringe(tmp_path: Path) -> None:
    svc = _service(_make_project(tmp_path))
    df = next(iter(svc._load_traces("groom").values()))
    bouts = svc._detect_bouts("groom", df)
    # Two separated bouts (merge_gap=1 keeps them apart).
    assert len(bouts) == 2
    means = sorted(round(m, 2) for _s, _e, m in bouts)
    assert means == [0.62, 0.9]


def test_negative_runs_detected(tmp_path: Path) -> None:
    svc = _service(_make_project(tmp_path))
    df = next(iter(svc._load_traces("groom").values()))
    negs = svc._detect_negative_runs("groom", df)
    # Baseline 0.1 regions before and after the bouts.
    assert len(negs) >= 1
    assert all(m < 0.4 for _s, _e, m in negs)


def test_sample_positive_excludes_fringe(tmp_path: Path) -> None:
    svc = _service(_make_project(tmp_path))
    settings = ValidationSettings()
    active = [("groom", "Grooming")]
    pos = svc._sample_positive(5, active, settings, fringe=False)
    assert len(pos) == 1
    assert pos[0].category == "unreviewed_positive"
    assert pos[0].machine_label == "groom"
    assert pos[0].is_fringe is False

    fringe = svc._sample_positive(5, active, settings, fringe=True)
    assert len(fringe) == 1
    assert fringe[0].category == "fringe"
    assert fringe[0].is_fringe is True


def test_sample_negative_labels_no_behavior(tmp_path: Path) -> None:
    svc = _service(_make_project(tmp_path))
    negs = svc._sample_negative(3, [("groom", "Grooming")], ValidationSettings())
    assert negs
    assert all(r.machine_label == "no_behavior" for r in negs)
    assert all(r.category == "negative" for r in negs)


# ---------------------------------------------------------------------------
# Metrics
# ---------------------------------------------------------------------------
def _clip(cid: str, machine: str, behavior: str = "groom", category: str = "unreviewed_positive") -> ValidationClipRecord:
    return ValidationClipRecord(
        clip_id=cid,
        category=category,  # type: ignore[arg-type]
        behavior_id=behavior,
        machine_label=machine,
        session_id="s",
        start_frame=0,
        end_frame=10,
    )


def test_compute_metrics_user_vs_machine(tmp_path: Path) -> None:
    svc = ValidationService()
    run = ValidationRun(
        run_id="r1",
        clips=[
            _clip("c1", machine="groom"),
            _clip("c2", machine="groom"),
            _clip("c3", machine="no_behavior", category="negative"),
            _clip("c4", machine="no_behavior", category="negative"),
        ],
    )
    # Reviewer agrees on c1, c3, c4; disagrees on c2 (model said groom, human says negative).
    answers = {
        "rev1": {
            "c1": ValidationAnswerRecord(clip_id="c1", reviewer_id="rev1", label="groom"),
            "c2": ValidationAnswerRecord(clip_id="c2", reviewer_id="rev1", label="no_behavior"),
            "c3": ValidationAnswerRecord(clip_id="c3", reviewer_id="rev1", label="no_behavior"),
            "c4": ValidationAnswerRecord(clip_id="c4", reviewer_id="rev1", label="no_behavior"),
        }
    }
    metrics = svc.compute_metrics(run, answers)
    rdata = metrics["per_reviewer"]["rev1"]
    assert rdata["n_answered"] == 4
    assert rdata["agreement"] == 0.75
    pb = rdata["per_behavior"]["groom"]
    # Model predicted groom for c1,c2; human agreed on c1 only → TP=1, FP=1, FN=0.
    assert pb["tp"] == 1
    assert pb["fp"] == 1
    assert pb["fn"] == 0
    assert pb["precision"] == 0.5
    assert pb["recall"] == 1.0


def test_unsure_excluded_from_precision(tmp_path: Path) -> None:
    svc = ValidationService()
    run = ValidationRun(run_id="r", clips=[_clip("c1", "groom")])
    answers = {
        "rev1": {"c1": ValidationAnswerRecord(clip_id="c1", reviewer_id="rev1", label="no_behavior", is_unsure=True)}
    }
    metrics = svc.compute_metrics(run, answers)
    rdata = metrics["per_reviewer"]["rev1"]
    assert rdata["n_answered"] == 0
    assert rdata["n_unsure"] == 1
    assert rdata["unsure_rate"] == 1.0


# ---------------------------------------------------------------------------
# Inter-rater kappa
# ---------------------------------------------------------------------------
def test_inter_rater_perfect_agreement(tmp_path: Path) -> None:
    svc = ValidationService()
    run = ValidationRun(
        run_id="r",
        clips=[_clip("c1", "groom"), _clip("c2", "no_behavior", category="negative")],
    )
    answers = {
        "a": {
            "c1": ValidationAnswerRecord(clip_id="c1", reviewer_id="a", label="groom"),
            "c2": ValidationAnswerRecord(clip_id="c2", reviewer_id="a", label="no_behavior"),
        },
        "b": {
            "c1": ValidationAnswerRecord(clip_id="c1", reviewer_id="b", label="groom"),
            "c2": ValidationAnswerRecord(clip_id="c2", reviewer_id="b", label="no_behavior"),
        },
    }
    inter = svc.compute_metrics(run, answers)["inter_rater"]
    assert inter["n_reviewers"] == 2
    assert inter["shared_clips"] == 2
    assert inter["agreement"] == 1.0
    assert inter["kappa"] == 1.0


def test_fleiss_kappa_three_reviewers(tmp_path: Path) -> None:
    svc = ValidationService()
    clips = [_clip(f"c{i}", "groom" if i % 2 == 0 else "no_behavior") for i in range(4)]
    run = ValidationRun(run_id="r", clips=clips)

    def _ans(rev: str, mapping: dict[str, str]) -> dict[str, ValidationAnswerRecord]:
        return {cid: ValidationAnswerRecord(clip_id=cid, reviewer_id=rev, label=lab) for cid, lab in mapping.items()}

    truth = {c.clip_id: c.machine_label for c in clips}
    answers = {
        "a": _ans("a", truth),
        "b": _ans("b", truth),
        "c": _ans("c", truth),
    }
    inter = svc.compute_metrics(run, answers)["inter_rater"]
    assert inter["n_reviewers"] == 3
    assert inter["kappa"] == 1.0  # all three agree on every clip


# ---------------------------------------------------------------------------
# Settings round-trip
# ---------------------------------------------------------------------------
def test_model_overview_resolves_per_behavior_dir(tmp_path: Path) -> None:
    root = _make_project(tmp_path)
    mdir = root / "derived" / "models" / "behavior_model_Grooming"
    mdir.mkdir(parents=True)
    (mdir / "metrics.json").write_text(
        json.dumps({"f1": 0.77, "precision": 0.81, "recall": 0.73, "pr_auc": 0.85, "n_train": 100, "n_val": 25}),
        encoding="utf-8",
    )
    svc = _service(root)
    rows = svc.model_overview()
    by_name = {r["behavior_name"]: r for r in rows}
    # No Behavior is its own model row.
    assert "No Behavior" in by_name
    row = by_name["Grooming"]
    assert row["model_version"] == "behavior_model_Grooming"
    assert abs(row["frame_f1"] - 0.77) < 1e-9
    assert row["n_train"] == 100
    # Bouts come from the competition trace; the clear bout should be counted.
    assert row["n_bouts"] >= 1


def test_delete_run_removes_run_answers_and_active_pointer(tmp_path: Path) -> None:
    svc = _service(_make_project(tmp_path))
    run = ValidationRun(run_id="del1", clips=[_clip("c1", "groom")])
    svc._save_run(run)
    from abel.storage.file_store import write_json
    write_json(svc._active_run_path(), {"run_id": "del1"})
    svc.save_answer("del1", ValidationAnswerRecord(clip_id="c1", reviewer_id="alice", label="groom"))

    assert svc.load_run("del1") is not None
    assert svc.list_reviewers("del1") == ["alice"]

    assert svc.delete_run("del1") is True
    assert svc.load_run("del1") is None
    assert svc.list_reviewers("del1") == []
    assert svc.load_active_run() is None
    # Idempotent: deleting again is a no-op, not an error.
    assert svc.delete_run("del1") is False


def test_prune_orphan_clips_keeps_referenced_only(tmp_path: Path) -> None:
    svc = _service(_make_project(tmp_path))
    clips_dir = svc._clips_dir() / "session_a"
    clips_dir.mkdir(parents=True)
    keep = clips_dir / "keep.mp4"
    orphan = clips_dir / "orphan.mp4"
    keep.write_bytes(b"x")
    orphan.write_bytes(b"x")

    run = ValidationRun(
        run_id="r",
        clips=[ValidationClipRecord(
            clip_id="c1", category="negative", behavior_id="groom", machine_label="no_behavior",
            session_id="session_a", start_frame=0, end_frame=5, clip_path=str(keep),
        )],
    )
    svc._save_run(run)

    svc._prune_orphan_clips()
    assert keep.exists()        # referenced by a saved run
    assert not orphan.exists()  # stale/unreferenced -> removed


def test_existing_clip_reuse_only_trusts_review_clips(tmp_path: Path) -> None:
    svc = _service(_make_project(tmp_path))
    sid = "session_a"
    from abel.services.preprocessing_service import ClipExtractionService
    stem = ClipExtractionService.clip_filename_for_id("val_negative_groom_session_a_0_5")

    # A stale clip in the validation cache must NOT be reused...
    val_cache = svc._clips_dir() / sid
    val_cache.mkdir(parents=True)
    (val_cache / f"{stem}.mp4").write_bytes(b"x")
    assert svc._find_existing_clip("val_negative_groom_session_a_0_5", sid) is None

    # ...but a subject-centered Clip Review clip is reused.
    review = svc._root() / "derived" / "clips" / sid
    review.mkdir(parents=True)
    review_clip = review / f"{stem}.mp4"
    review_clip.write_bytes(b"x")
    assert svc._find_existing_clip("val_negative_groom_session_a_0_5", sid) == review_clip


def test_intra_rater_self_consistency(tmp_path: Path) -> None:
    svc = ValidationService()
    # Prior-accepted clips carry the reviewer's original label as reference_label.
    clips = [
        ValidationClipRecord(
            clip_id=f"p{i}", category="prior_accepted", behavior_id="groom",
            machine_label="groom", reference_label="groom",
            session_id="s", start_frame=0, end_frame=5,
        )
        for i in range(4)
    ]
    # A non-prior clip should be ignored by intra-rater.
    clips.append(_clip("x1", machine="groom"))
    run = ValidationRun(run_id="r", clips=clips)
    # Reviewer re-labels 3/4 consistently, 1 differently, ignores the machine-only clip.
    answers = {
        "alice": {
            "p0": ValidationAnswerRecord(clip_id="p0", reviewer_id="alice", label="groom"),
            "p1": ValidationAnswerRecord(clip_id="p1", reviewer_id="alice", label="groom"),
            "p2": ValidationAnswerRecord(clip_id="p2", reviewer_id="alice", label="groom"),
            "p3": ValidationAnswerRecord(clip_id="p3", reviewer_id="alice", label="no_behavior"),
            "x1": ValidationAnswerRecord(clip_id="x1", reviewer_id="alice", label="no_behavior"),
        }
    }
    intra = svc.compute_metrics(run, answers)["intra_rater"]["alice"]
    assert intra["n"] == 4  # only prior-accepted clips counted
    assert intra["agreement"] == 0.75
    assert intra["per_behavior_counts"]["groom"] == 4


def test_overlap_clips_excluded_from_scoring(tmp_path: Path) -> None:
    svc = ValidationService()
    clips = [
        _clip("c1", machine="groom"),  # single behavior
        ValidationClipRecord(  # ambiguous: model flagged two behaviors at once
            clip_id="c2", category="unreviewed_positive", behavior_id="groom",
            machine_label="groom", session_id="s", start_frame=0, end_frame=5,
            coactive_labels=["groom", "rear"],
        ),
    ]
    run = ValidationRun(run_id="r", clips=clips)
    answers = {
        "alice": {
            "c1": ValidationAnswerRecord(clip_id="c1", reviewer_id="alice", label="groom"),
            # Picks the *other* flagged behavior on the overlap clip — must not be penalized.
            "c2": ValidationAnswerRecord(clip_id="c2", reviewer_id="alice", label="rear"),
        }
    }
    # No project set -> _assign_coactive_labels is a no-op, preserving the constructed labels.
    rdata = svc.compute_metrics(run, answers)["per_reviewer"]["alice"]
    assert rdata["n_answered"] == 1          # overlap clip excluded
    assert rdata["agreement"] == 1.0         # only the clean clip counts, and it agrees
    assert rdata["n_overlap"] == 1
    assert rdata["n_overlap_matched"] == 1   # picked a flagged behavior


def test_apply_inhibition_writes_symmetric_matrix(tmp_path: Path) -> None:
    svc = _service(_make_project(tmp_path))
    w = svc.apply_inhibition("groom", "rear", weight=0.25)
    assert w == 0.25
    import json
    cfg = json.loads((svc._root() / "config" / "temporal_refinement_settings.json").read_text())
    matrix = cfg["__all__"]["suppression_matrix"]
    assert matrix["groom"]["rear"] == 0.25
    assert matrix["rear"]["groom"] == 0.25  # symmetric / mutual


def test_excluded_behaviors_dropped_from_competition(tmp_path: Path) -> None:
    root = _make_project(tmp_path)
    # Add a second behavior that is NOT part of the competition behavior_models.
    import yaml
    bpath = root / "config" / "behavior_definitions.yaml"
    data = yaml.safe_load(bpath.read_text())
    data["behaviors"].append({
        "behavior_id": "rear", "name": "Rear", "short_name": "rear", "is_active": True,
    })
    bpath.write_text(yaml.safe_dump(data))
    # Competition manifest lists only "groom".
    import json
    inf_dir = root / "derived" / "inference"
    man = json.loads((inf_dir / "inference_manifest.json").read_text())
    man["competition"] = {"behavior_models": {"groom": "behavior_model_Grooming"},
                          "excluded_behavior_ids": ["rear"]}
    (inf_dir / "inference_manifest.json").write_text(json.dumps(man))
    # Competition runs live under the target_behavior token.
    tb_dir = root / "derived" / "temporal_refinement" / "target_behavior"
    tb_dir.mkdir(parents=True, exist_ok=True)
    (tb_dir / "latest.json").write_text(
        json.dumps({"inference_dir": str(inf_dir), "postprocess_dir": ""})
    )

    svc = _service(root)
    names = [n for _, n in svc._active_behaviors()]
    assert names == ["Grooming"]  # Rear excluded from competition


def test_export_results_xlsx(tmp_path: Path) -> None:
    svc = _service(_make_project(tmp_path))
    run = ValidationRun(run_id="r1", clips=[_clip("c1", "groom")])
    svc._save_run(run)
    svc.save_answer("r1", ValidationAnswerRecord(clip_id="c1", reviewer_id="alice", label="groom"))
    out = tmp_path / "out.xlsx"
    n = svc.export_results_xlsx(out)
    assert n == 1
    assert out.exists() and out.stat().st_size > 0
    import openpyxl
    wb = openpyxl.load_workbook(out)
    assert "Index" in wb.sheetnames
    assert any(s not in ("Index", "Model Summary") for s in wb.sheetnames)


def test_settings_roundtrip(tmp_path: Path) -> None:
    svc = _service(_make_project(tmp_path))
    s = svc.load_settings()
    s.n_total_clips = 42
    s.prop_fringe = 0.2
    svc.save_settings(s)
    reloaded = svc.load_settings()
    assert reloaded.n_total_clips == 42
    assert reloaded.prop_fringe == 0.2


def test_confusion_analysis_tracks_pairs_and_fringe(tmp_path: Path) -> None:
    svc = ValidationService()
    run = ValidationRun(
        run_id="r",
        clips=[
            _clip("c1", machine="groom"),  # model: groom
            _clip("c2", machine="groom"),  # model: groom
            ValidationClipRecord(
                clip_id="c3", category="fringe", behavior_id="groom",
                machine_label="groom", session_id="s", start_frame=0, end_frame=5, is_fringe=True,
            ),
        ],
    )
    answers = {
        "rev1": {
            "c1": ValidationAnswerRecord(clip_id="c1", reviewer_id="rev1", label="groom"),       # agree
            "c2": ValidationAnswerRecord(clip_id="c2", reviewer_id="rev1", label="no_behavior"),  # clear disagree
            "c3": ValidationAnswerRecord(clip_id="c3", reviewer_id="rev1", label="no_behavior"),  # fringe disagree
        }
    }
    conf = svc.compute_metrics(run, answers)["confusion"]
    assert conf["n_disagreements"] == 2
    assert conf["n_fringe_disagreements"] == 1
    assert conf["n_clear_disagreements"] == 1
    # The (groom -> no_behavior) confusion should be the top pair with count 2, 1 fringe.
    top = conf["top_confusions"][0]
    assert top["machine"] == "groom"
    assert top["user"] == "no_behavior"
    assert top["count"] == 2
    assert top["fringe_count"] == 1


def test_suggestions_flags_low_precision(tmp_path: Path) -> None:
    svc = _service(_make_project(tmp_path))
    metrics = {
        "behaviors": ["groom"],
        "per_reviewer": {
            "rev1": {
                "unsure_rate": 0.0,
                "per_behavior": {"groom": {"precision": 0.3, "recall": 0.9, "f1": 0.45}},
            }
        },
        "inter_rater": {"n_reviewers": 1, "kappa": None},
    }
    suggestions = svc.suggestions(metrics)
    assert any("over-predicts" in s["message"] for s in suggestions)
